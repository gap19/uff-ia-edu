"""Conversão de tabelas TIC Educação 2023 (Excel) para Parquet.

Lê os 4 arquivos XLSX da pesquisa TIC Educação 2023 (proporção, total,
margem de erro e margem de erro total), normaliza as planilhas em
formato longo e exporta um único Parquet consolidado.

Cada planilha representa um indicador (ex.: A1, B3) e possui cabeçalhos
multi-nível com células mescladas.  O parser tenta extrair a estrutura
de forma robusta, pulando planilhas que não possam ser interpretadas.
"""

from __future__ import annotations

import logging
import re
import warnings
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from backend.config import TIC_PARQUET_DIR, TIC_XLSX_FILES

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

# Número máximo de linhas de cabeçalho a inspecionar
_MAX_HEADER_ROWS = 8

# Padrão para detectar linhas de "variável de corte" (ex.: "TOTAL", "REGIÃO")
_VARIAVEL_CORTE_RE = re.compile(
    r"^(TOTAL|REGIÃO|DEPENDÊNCIA|ÁREA|PORTE|SÉRIE|RENDA|SEXO|FAIXA|IDADE|"
    r"ESCOLARIDADE|CLASSE|TIPO|LOCALIZAÇÃO|NÍVEL|GRAU|MODALIDADE|TURNO|"
    r"ESFERA|CONDIÇÃO|ESTADO|UF|GRANDE REGIÃO)",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _cell_value(cell: Cell | MergedCell | None) -> Any:
    """Retorna o valor de uma célula, tratando MergedCell como None."""
    if cell is None:
        return None
    return cell.value


def _clean_text(value: Any) -> str | None:
    """Converte valor de célula em texto limpo ou None."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text


def _parse_numeric(value: Any) -> float | None:
    """Tenta converter um valor de célula em float.

    Retorna None para valores não-numéricos como '-', '*', '..', texto vazio.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    # Valores comuns que indicam ausência de dado
    if text in ("", "-", "*", "..", "…", "—", "–", "0*", "**"):
        return None
    # Tenta converter separador decimal brasileiro
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _resolve_merged_value(ws: Worksheet, row: int, col: int) -> Any:
    """Resolve o valor real de uma célula, mesmo se estiver em região mesclada."""
    cell = ws.cell(row=row, column=col)
    val = cell.value
    if val is not None:
        return val
    # Procura em merged_cells se esta célula pertence a algum range
    # (read_only worksheets não possuem merged_cells)
    merged = getattr(ws, "merged_cells", None)
    if merged is not None:
        for merged_range in merged.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                ).value
    return None


def _find_header_boundary(ws: Worksheet) -> int:
    """Detecta onde terminam os cabeçalhos e começam os dados.

    Retorna o número da primeira linha de dados (1-indexed).
    Heurística: a primeira linha cujo conteúdo na coluna A case com um
    padrão de variável de corte, ou a primeira linha após as _MAX_HEADER_ROWS
    iniciais que tenha conteúdo na coluna A.
    """
    for row_idx in range(1, min(ws.max_row or 1, _MAX_HEADER_ROWS + 1) + 1):
        text = _clean_text(_resolve_merged_value(ws, row_idx, 1))
        if text and _VARIAVEL_CORTE_RE.match(text):
            return row_idx
    # Fallback: assume que os dados começam na linha 5
    return 5


def _build_column_headers(
    ws: Worksheet, data_start_row: int
) -> list[str]:
    """Constrói rótulos de coluna a partir das linhas de cabeçalho multi-nível.

    Combina as linhas 1..(data_start_row - 1) concatenando textos
    não-None com ' | '.
    """
    header_rows = range(1, data_start_row)
    max_col = ws.max_column or 1
    headers: list[str] = []

    for col_idx in range(1, max_col + 1):
        parts: list[str] = []
        for row_idx in header_rows:
            val = _clean_text(_resolve_merged_value(ws, row_idx, col_idx))
            if val and val not in parts:
                parts.append(val)
        header = " | ".join(parts) if parts else f"col_{col_idx}"
        headers.append(header)

    return headers


# ---------------------------------------------------------------------------
# Parser principal de planilha
# ---------------------------------------------------------------------------


def parse_tic_sheet(
    ws: Worksheet,
    sheet_name: str,
    tipo: str,
) -> list[dict[str, Any]]:
    """Analisa uma planilha TIC e retorna linhas normalizadas.

    Parameters
    ----------
    ws:
        Planilha openpyxl já carregada.
    sheet_name:
        Nome da planilha (código do indicador, ex.: "A1").
    tipo:
        Tipo de arquivo: 'proporcao', 'total', 'margem_erro',
        'margem_erro_total'.

    Returns
    -------
    list[dict]
        Lista de dicts com chaves: indicador, tipo, variavel_corte,
        valor_corte, regiao, valor.
    """
    rows: list[dict[str, Any]] = []

    if (ws.max_row or 0) < 3 or (ws.max_column or 0) < 2:
        logger.warning(
            "[tic_loader] Planilha '%s' vazia ou muito pequena, pulando.",
            sheet_name,
        )
        return rows

    # Detecta limite cabeçalho / dados
    data_start = _find_header_boundary(ws)
    col_headers = _build_column_headers(ws, data_start)
    max_col = len(col_headers)

    # Variável de corte corrente (grupo de linhas)
    current_variavel_corte: str | None = None

    for row_idx in range(data_start, (ws.max_row or data_start) + 1):
        first_cell = _clean_text(_resolve_merged_value(ws, row_idx, 1))

        if first_cell is None:
            continue

        # Detecta se é uma linha de cabeçalho de variável de corte
        if _VARIAVEL_CORTE_RE.match(first_cell):
            current_variavel_corte = first_cell
            # Verifica se a mesma linha também contém dados (a partir da col 2)
            has_data = False
            for col_idx in range(2, max_col + 1):
                val = _parse_numeric(
                    _resolve_merged_value(ws, row_idx, col_idx)
                )
                if val is not None:
                    has_data = True
                    break

            if not has_data:
                continue
            # Se tem dados, trata como linha de dados também (ex.: "TOTAL")

        valor_corte = first_cell

        # Itera pelas colunas de dados
        for col_idx in range(2, max_col + 1):
            raw_val = _resolve_merged_value(ws, row_idx, col_idx)
            numeric_val = _parse_numeric(raw_val)
            if numeric_val is None:
                continue

            col_header = col_headers[col_idx - 1] if col_idx - 1 < len(col_headers) else f"col_{col_idx}"

            rows.append(
                {
                    "indicador": sheet_name,
                    "tipo": tipo,
                    "variavel_corte": current_variavel_corte or "TOTAL",
                    "valor_corte": valor_corte,
                    "regiao": col_header,
                    "valor": numeric_val,
                }
            )

    return rows


# ---------------------------------------------------------------------------
# Inventário de indicadores
# ---------------------------------------------------------------------------


def get_tic_indicators() -> list[str]:
    """Retorna lista de códigos de indicadores disponíveis nos arquivos TIC.

    Lê os nomes das planilhas do primeiro arquivo disponível.
    """
    for tipo, path in TIC_XLSX_FILES.items():
        if path.exists():
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            names = list(wb.sheetnames)
            wb.close()
            return names
    return []


# ---------------------------------------------------------------------------
# Carga principal
# ---------------------------------------------------------------------------


def _open_workbook_safe(xlsx_path: Path):
    """Abre workbook ignorando drawings corrompidos.

    Alguns XLSX da pesquisa TIC referenciam 'xl/drawings/drawing1.xml' que
    não existe no arquivo ZIP.  O openpyxl tenta ler esse item e falha com
    KeyError.  Aqui fazemos monkey-patch temporário no ZipFile.read para
    retornar XML vazio em vez de falhar.
    """
    import zipfile

    _orig_read = zipfile.ZipFile.read

    def _patched_read(self, name, pwd=None):
        try:
            return _orig_read(self, name, pwd)
        except KeyError:
            if "drawing" in str(name).lower():
                logger.debug(
                    "[tic_loader] Drawing ausente ignorado: %s", name
                )
                return b'<?xml version="1.0" encoding="UTF-8"?><root/>'
            raise

    zipfile.ZipFile.read = _patched_read
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(
                str(xlsx_path), read_only=False, data_only=True
            )
    finally:
        zipfile.ZipFile.read = _orig_read

    return wb


def load_all_tic() -> int:
    """Processa todos os 4 arquivos TIC e exporta Parquet consolidado.

    Returns
    -------
    int
        Total de linhas no Parquet resultante.
    """
    TIC_PARQUET_DIR.mkdir(parents=True, exist_ok=True)
    output_path = TIC_PARQUET_DIR / "tic_educacao_2023.parquet"

    all_rows: list[dict[str, Any]] = []
    total_sheets = 0
    skipped_sheets = 0

    for tipo, xlsx_path in TIC_XLSX_FILES.items():
        if not xlsx_path.exists():
            logger.warning(
                "[tic_loader] Arquivo não encontrado: %s, pulando.", xlsx_path
            )
            continue

        print(f"[tic_loader] Abrindo {xlsx_path.name} (tipo={tipo})...")

        # read_only=False para resolver merged cells; mas alguns XLSX
        # têm referências a drawings corrompidos (bug conhecido openpyxl).
        # Patch temporário: intercepta KeyError em leitura de drawings.
        try:
            wb = _open_workbook_safe(xlsx_path)
        except Exception as exc:
            logger.error(
                "[tic_loader] Falha ao abrir %s: %s", xlsx_path.name, exc
            )
            continue

        for sheet_name in wb.sheetnames:
            total_sheets += 1
            try:
                ws = wb[sheet_name]
                sheet_rows = parse_tic_sheet(ws, sheet_name, tipo)
                all_rows.extend(sheet_rows)
                print(
                    f"  [{tipo}] {sheet_name}: {len(sheet_rows)} registros"
                )
            except Exception:
                skipped_sheets += 1
                logger.warning(
                    "[tic_loader] Erro ao processar planilha '%s' de %s, pulando.",
                    sheet_name,
                    xlsx_path.name,
                    exc_info=True,
                )

        wb.close()

    if not all_rows:
        warnings.warn(
            "[tic_loader] Nenhum dado extraído dos arquivos TIC.",
            stacklevel=2,
        )
        return 0

    # --- Salvar como Parquet via pyarrow ---
    print(f"\n[tic_loader] Total de registros: {len(all_rows):,}")
    print(f"[tic_loader] Planilhas processadas: {total_sheets - skipped_sheets}/{total_sheets}")

    _save_parquet(all_rows, output_path)

    print(f"[tic_loader] Parquet salvo em: {output_path}")
    return len(all_rows)


def _save_parquet(rows: list[dict[str, Any]], path: Path) -> None:
    """Salva lista de dicts como Parquet usando pyarrow."""
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq

        # Organiza por colunas
        columns = {
            "indicador": [r["indicador"] for r in rows],
            "tipo": [r["tipo"] for r in rows],
            "variavel_corte": [r["variavel_corte"] for r in rows],
            "valor_corte": [r["valor_corte"] for r in rows],
            "regiao": [r["regiao"] for r in rows],
            "valor": [r["valor"] for r in rows],
        }

        table = pa.table(
            {
                "indicador": pa.array(columns["indicador"], type=pa.string()),
                "tipo": pa.array(columns["tipo"], type=pa.string()),
                "variavel_corte": pa.array(columns["variavel_corte"], type=pa.string()),
                "valor_corte": pa.array(columns["valor_corte"], type=pa.string()),
                "regiao": pa.array(columns["regiao"], type=pa.string()),
                "valor": pa.array(columns["valor"], type=pa.float64()),
            }
        )

        pq.write_table(table, str(path), compression="snappy")

    except ImportError:
        # Fallback: usa duckdb para escrever Parquet
        import duckdb

        con = duckdb.connect()
        try:
            con.execute("""
                CREATE TABLE tic_data (
                    indicador VARCHAR,
                    tipo VARCHAR,
                    variavel_corte VARCHAR,
                    valor_corte VARCHAR,
                    regiao VARCHAR,
                    valor DOUBLE
                )
            """)

            con.executemany(
                "INSERT INTO tic_data VALUES (?, ?, ?, ?, ?, ?)",
                [
                    (
                        r["indicador"],
                        r["tipo"],
                        r["variavel_corte"],
                        r["valor_corte"],
                        r["regiao"],
                        r["valor"],
                    )
                    for r in rows
                ],
            )

            con.execute(
                f"COPY tic_data TO '{path}' (FORMAT PARQUET, COMPRESSION SNAPPY)"
            )
        finally:
            con.close()


# ---------------------------------------------------------------------------
# Execução direta
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.WARNING)
    total = load_all_tic()
    print(f"\n[tic_loader] Concluído: {total:,} registros totais.")
